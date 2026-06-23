import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

TEMPLATE_FINAL_PATH = None  # Can be set by caller before process_files()

def read_summary5(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    def v(row, col):
        val = ws.cell(row=row, column=col).value
        if val is None or val == '':
            return 0
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0

    def s(row, col):
        val = ws.cell(row=row, column=col).value
        return str(val) if val is not None else ''

    info = {
        'code': s(2, 5), 'record_no': s(2, 14),
        'site_name': s(3, 5), 'location_type': s(3, 14),
        'org_request': s(4, 5), 'responsible': s(4, 14),
        'store_code': s(5, 5), 'location_desc': s(5, 14),
        'org_unit': s(6, 5), 'province': s(6, 14),
        'count_type': s(7, 5), 'backup_rank': s(8, 6),
        'day1': s(10, 7), 'day2': s(10, 12),
    }

    data = {}
    # คน
    data['male']          = {'d1': [v(12,7),v(12,8),v(12,9),v(12,10)], 'd2': [v(12,12),v(12,13),v(12,14),v(12,15)]}
    data['female']        = {'d1': [v(13,7),v(13,8),v(13,9),v(13,10)], 'd2': [v(13,12),v(13,13),v(13,14),v(13,15)]}
    data['child']         = {'d1': [v(14,7),v(14,8),v(14,9),v(14,10)], 'd2': [v(14,12),v(14,13),v(14,14),v(14,15)]}
    data['teen']          = {'d1': [v(15,7),v(15,8),v(15,9),v(15,10)], 'd2': [v(15,12),v(15,13),v(15,14),v(15,15)]}
    data['working']       = {'d1': [v(16,7),v(16,8),v(16,9),v(16,10)], 'd2': [v(16,12),v(16,13),v(16,14),v(16,15)]}
    data['adult']         = {'d1': [v(17,7),v(17,8),v(17,9),v(17,10)], 'd2': [v(17,12),v(17,13),v(17,14),v(17,15)]}
    data['student_elem']  = {'d1': [v(18,7),v(18,8),v(18,9),v(18,10)], 'd2': [v(18,12),v(18,13),v(18,14),v(18,15)]}
    data['student_uni']   = {'d1': [v(19,7),v(19,8),v(19,9),v(19,10)], 'd2': [v(19,12),v(19,13),v(19,14),v(19,15)]}
    data['employee']      = {'d1': [v(20,7),v(20,8),v(20,9),v(20,10)], 'd2': [v(20,12),v(20,13),v(20,14),v(20,15)]}
    data['factory']       = {'d1': [v(21,7),v(21,8),v(21,9),v(21,10)], 'd2': [v(21,12),v(21,13),v(21,14),v(21,15)]}
    data['merchant']      = {'d1': [v(22,7),v(22,8),v(22,9),v(22,10)], 'd2': [v(22,12),v(22,13),v(22,14),v(22,15)]}
    data['tourist']       = {'d1': [v(23,7),v(23,8),v(23,9),v(23,10)], 'd2': [v(23,12),v(23,13),v(23,14),v(23,15)]}
    data['housewife']     = {'d1': [v(24,7),v(24,8),v(24,9),v(24,10)], 'd2': [v(24,12),v(24,13),v(24,14),v(24,15)]}
    data['other_person']  = {'d1': [v(25,7),v(25,8),v(25,9),v(25,10)], 'd2': [v(25,12),v(25,13),v(25,14),v(25,15)]}
    data['dir_left']      = {'d1': [v(26,7),v(26,8),v(26,9),v(26,10)], 'd2': [v(26,12),v(26,13),v(26,14),v(26,15)]}
    data['dir_right']     = {'d1': [v(27,7),v(27,8),v(27,9),v(27,10)], 'd2': [v(27,12),v(27,13),v(27,14),v(27,15)]}
    data['total_people']  = {'d1': [v(28,7),v(28,8),v(28,9),v(28,10)], 'd2': [v(28,12),v(28,13),v(28,14),v(28,15)]}
    # รถ
    data['front_moto']    = {'d1': [v(29,7),v(29,8),v(29,9),v(29,10)], 'd2': [v(29,12),v(29,13),v(29,14),v(29,15)]}
    data['front_car']     = {'d1': [v(30,7),v(30,8),v(30,9),v(30,10)], 'd2': [v(30,12),v(30,13),v(30,14),v(30,15)]}
    data['front_total']   = {'d1': [v(31,7),v(31,8),v(31,9),v(31,10)], 'd2': [v(31,12),v(31,13),v(31,14),v(31,15)]}
    data['opp_moto']      = {'d1': [v(32,7),v(32,8),v(32,9),v(32,10)], 'd2': [v(32,12),v(32,13),v(32,14),v(32,15)]}
    data['opp_car']       = {'d1': [v(33,7),v(33,8),v(33,9),v(33,10)], 'd2': [v(33,12),v(33,13),v(33,14),v(33,15)]}
    data['opp_total']     = {'d1': [v(34,7),v(34,8),v(34,9),v(34,10)], 'd2': [v(34,12),v(34,13),v(34,14),v(34,15)]}
    data['all_moto']      = {'d1': [v(35,7),v(35,8),v(35,9),v(35,10)], 'd2': [v(35,12),v(35,13),v(35,14),v(35,15)]}
    data['all_car']       = {'d1': [v(36,7),v(36,8),v(36,9),v(36,10)], 'd2': [v(36,12),v(36,13),v(36,14),v(36,15)]}
    data['all_total']     = {'d1': [v(37,7),v(37,8),v(37,9),v(37,10)], 'd2': [v(37,12),v(37,13),v(37,14),v(37,15)]}

    for key in data:
        d = data[key]
        d['avg'] = [(d['d1'][i] + d['d2'][i]) / 2 for i in range(4)]

    tp = data['total_people']['avg'][3]
    tt = data['all_total']['avg'][3]
    ft = data['front_total']['avg'][3]
    ot = data['opp_total']['avg'][3]

    def pct(key, denom):
        return data[key]['avg'][3] / denom if denom else 0

    data['pct'] = {
        'male': pct('male', tp), 'female': pct('female', tp),
        'child': pct('child', tp), 'teen': pct('teen', tp),
        'working': pct('working', tp), 'adult': pct('adult', tp),
        'student_elem': pct('student_elem', tp), 'student_uni': pct('student_uni', tp),
        'employee': pct('employee', tp), 'factory': pct('factory', tp),
        'merchant': pct('merchant', tp), 'tourist': pct('tourist', tp),
        'housewife': pct('housewife', tp), 'other_person': pct('other_person', tp),
        'dir_left': pct('dir_left', tp), 'dir_right': pct('dir_right', tp),
        'front_moto': pct('front_moto', ft), 'front_car': pct('front_car', ft),
        'opp_moto': pct('opp_moto', ot), 'opp_car': pct('opp_car', ot),
        'all_moto': pct('all_moto', tt), 'all_car': pct('all_car', tt),
        'front_total': pct('front_total', tt), 'opp_total': pct('opp_total', tt),
    }
    return info, data


def build_summary1(info, data, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    p = data['pct']
    a = lambda key, i: data[key]['avg'][i]

    def num(v):
        return int(v) if v == int(v) else v

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12

    thin = Side(style='thin')
    def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)
    def fnt(bold=False): return Font(name='TH SarabunPSK', size=14, bold=bold)
    def aln(h='center'): return Alignment(horizontal=h, vertical='center', wrap_text=True)

    def hdr(row, col, val, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = fnt(bold); c.alignment = aln()

    def row_data(r, c1, c2, c3, m, e, n, t, pct_val):
        ws.row_dimensions[r].height = 18
        if c1: ws.cell(row=r, column=1, value=c1).font = fnt(True); ws.cell(row=r, column=1).alignment = aln()
        if c2: ws.cell(row=r, column=2, value=c2).font = fnt(True); ws.cell(row=r, column=2).alignment = aln()
        if c3: ws.cell(row=r, column=3, value=c3).font = fnt(); ws.cell(row=r, column=3).alignment = aln('left')
        for col, val in [(4, m), (5, e), (6, n), (7, t)]:
            c = ws.cell(row=r, column=col, value=num(val) if val is not None else '')
            c.font = fnt(); c.alignment = aln(); c.border = bdr()
        c = ws.cell(row=r, column=8, value=pct_val if pct_val != '' else None)
        c.font = fnt(); c.alignment = aln(); c.border = bdr()
        if pct_val != '': c.number_format = '0.00%'

    hdr(2, 1, 'หัวข้อ', True); ws.merge_cells('A2:C2')
    hdr(2, 4, 'เฉลี่ย', True); ws.merge_cells('D2:H2')
    for col, lbl in [(4,'ผลัดเช้า'),(5,'ผลัดบ่าย'),(6,'ผลัดดึก'),(7,'รวมทั้งวัน'),(8,'%สัดส่วน')]:
        hdr(3, col, lbl, True)

    row_data(4,  'คน','เพศ','ผู้ชาย',           a('male',0),a('male',1),a('male',2),a('male',3),p['male'])
    row_data(5,  None,None,'ผู้หญิง',            a('female',0),a('female',1),a('female',2),a('female',3),p['female'])
    row_data(6,  None,'อายุ','วัยเด็ก(อายุ 4-12 ปี)', a('child',0),a('child',1),a('child',2),a('child',3),p['child'])
    row_data(7,  None,None,'วัยรุ่น (13-22 ปี)', a('teen',0),a('teen',1),a('teen',2),a('teen',3),p['teen'])
    row_data(8,  None,None,'วัยทำงาน (23-34 ปี)',a('working',0),a('working',1),a('working',2),a('working',3),p['working'])
    row_data(9,  None,None,'วัยผู้ใหญ่ (35 ปีขี้นไป)',a('adult',0),a('adult',1),a('adult',2),a('adult',3),p['adult'])
    row_data(10, None,'อาชีพ','นร.อนุบาล-ประถม',  a('student_elem',0),a('student_elem',1),a('student_elem',2),a('student_elem',3),p['student_elem'])
    row_data(11, None,None,'นร.ม.ต้น-มหาวิทยาลัย',a('student_uni',0),a('student_uni',1),a('student_uni',2),a('student_uni',3),p['student_uni'])
    row_data(12, None,None,'พนง.เอกชน-ข้าราชการ', a('employee',0),a('employee',1),a('employee',2),a('employee',3),p['employee'])
    row_data(13, None,None,'คนโรงงาน/รปภ./คนก่อสร้าง',a('factory',0),a('factory',1),a('factory',2),a('factory',3),p['factory'])
    row_data(14, None,None,'พ่อค้า/แม่ค้า',       a('merchant',0),a('merchant',1),a('merchant',2),a('merchant',3),p['merchant'])
    row_data(15, None,None,'นักท่องเที่ยว',        a('tourist',0),a('tourist',1),a('tourist',2),a('tourist',3),p['tourist'])
    row_data(16, None,None,'พ่อบ้านแม่บ้าน',       a('housewife',0),a('housewife',1),a('housewife',2),a('housewife',3),p['housewife'])
    row_data(17, None,None,'อื่นๆ(ขาจร)',          a('other_person',0),a('other_person',1),a('other_person',2),a('other_person',3),p['other_person'])
    row_data(18, None,'ทิศ','ซ้าย',               a('dir_left',0),a('dir_left',1),a('dir_left',2),a('dir_left',3),p['dir_left'])
    row_data(19, None,None,'ขวา',                  a('dir_right',0),a('dir_right',1),a('dir_right',2),a('dir_right',3),p['dir_right'])
    row_data(20, None,'จำนวนคน (ยอดรวม)',None,    a('total_people',0),a('total_people',1),a('total_people',2),a('total_people',3),'')
    ws.merge_cells('B20:C20')

    ws.merge_cells('A4:A20');  ws.cell(4,1).alignment = aln()
    ws.merge_cells('B4:B5');   ws.cell(4,2).alignment = aln()
    ws.merge_cells('B6:B9');   ws.cell(6,2).alignment = aln()
    ws.merge_cells('B10:B17'); ws.cell(10,2).alignment = aln()
    ws.merge_cells('B18:B19'); ws.cell(18,2).alignment = aln()

    row_data(21, 'รถ','ประเภทรถ','รถจย.+จยย.',           a('all_moto',0),a('all_moto',1),a('all_moto',2),a('all_moto',3),p['all_moto'])
    row_data(22, None,None,'รถอื่นๆ',                      a('all_car',0),a('all_car',1),a('all_car',2),a('all_car',3),p['all_car'])
    row_data(23, None,'ทิศ','รถผ่านหน้าฝั่งเป้าหมาย',    a('front_total',0),a('front_total',1),a('front_total',2),a('front_total',3),p['front_total'])
    row_data(24, None,None,'รถผ่านฝั่งตรงข้ามเป้าหมาย',  a('opp_total',0),a('opp_total',1),a('opp_total',2),a('opp_total',3),p['opp_total'])
    row_data(25, None,'จำนวนรถ (ยอดรวม)',None,            a('all_total',0),a('all_total',1),a('all_total',2),a('all_total',3),'')
    ws.merge_cells('B25:C25')

    ws.merge_cells('A21:A25'); ws.cell(21,1).alignment = aln()
    ws.merge_cells('B21:B22'); ws.cell(21,2).alignment = aln()
    ws.merge_cells('B23:B24'); ws.cell(23,2).alignment = aln()

    wb.save(out_path)


def build_summary_final(info, data, out_path, template_path=None):
    tp = template_path or TEMPLATE_FINAL_PATH or os.path.join(os.path.dirname(os.path.abspath(__file__)), 'template_final.xlsx')
    wb = openpyxl.load_workbook(tp)
    ws = wb.active
    p = data['pct']
    a = lambda key, i: data[key]['avg'][i]

    def ss(row, col, val):
        ws.cell(row=row, column=col).value = val

    ss(5,4,round(p['male'],4));       ss(6,4,round(p['female'],4))
    ss(8,4,round(p['child'],4));      ss(9,4,round(p['teen'],4))
    ss(10,4,round(p['working'],4));   ss(11,4,round(p['adult'],4))
    ss(13,4,round(p['student_elem'],4)); ss(14,4,round(p['student_uni'],4))
    ss(15,4,round(p['employee'],4));  ss(16,4,round(p['factory'],4))
    ss(17,4,round(p['merchant'],4));  ss(18,4,round(p['housewife'],4))
    ss(19,4,round(p['other_person'],4))
    ss(21,4,round(p['dir_left'],4));  ss(22,4,round(p['dir_right'],4))
    ss(23,4,round(a('total_people',3),0))
    ss(33,4,round(a('all_total',3),0))

    for row, idx in {51:0, 60:1, 69:2}.items():
        ss(row,3,a('total_people',idx)); ss(row,4,a('dir_left',idx)); ss(row,5,a('dir_right',idx))
        ss(row,6,a('male',idx));         ss(row,7,a('female',idx))
        ss(row,8,a('child',idx));        ss(row,9,a('teen',idx))
        ss(row,10,a('working',idx));     ss(row,11,a('adult',idx))
        ss(row,12,a('student_elem',idx)); ss(row,13,a('student_uni',idx))
        ss(row,14,a('employee',idx));    ss(row,15,a('factory',idx))
        ss(row,16,a('merchant',idx));    ss(row,17,a('housewife',idx))
        ss(row,18,a('other_person',idx)); ss(row,19,a('all_total',idx))
        ss(row,26,a('front_moto',idx));  ss(row,27,a('front_car',idx));  ss(row,28,a('front_total',idx))
        ss(row,29,a('opp_moto',idx));    ss(row,30,a('opp_car',idx));    ss(row,31,a('opp_total',idx))

    ss(70,3,a('total_people',3)); ss(70,4,a('dir_left',3)); ss(70,5,a('dir_right',3))
    ss(70,6,a('male',3));         ss(70,7,a('female',3))
    ss(70,8,a('child',3));        ss(70,9,a('teen',3))
    ss(70,10,a('working',3));     ss(70,11,a('adult',3))
    ss(70,12,a('student_elem',3)); ss(70,13,a('student_uni',3))
    ss(70,14,a('employee',3));    ss(70,15,a('factory',3))
    ss(70,16,a('merchant',3));    ss(70,17,a('housewife',3))
    ss(70,18,a('other_person',3)); ss(70,19,a('all_total',3))
    ss(70,26,a('front_moto',3));  ss(70,27,a('front_car',3));  ss(70,28,a('front_total',3))
    ss(70,29,a('opp_moto',3));    ss(70,30,a('opp_car',3));    ss(70,31,a('opp_total',3))

    ss(72,4,round(p['dir_left'],4));   ss(72,5,round(p['dir_right'],4))
    ss(72,6,round(p['male'],4));       ss(72,7,round(p['female'],4))
    ss(72,8,round(p['child'],4));      ss(72,9,round(p['teen'],4))
    ss(72,10,round(p['working'],4));   ss(72,11,round(p['adult'],4))
    ss(72,12,round(p['student_elem'],4)); ss(72,13,round(p['student_uni'],4))
    ss(72,14,round(p['employee'],4));  ss(72,15,round(p['factory'],4))
    ss(72,16,round(p['merchant'],4));  ss(72,17,round(p['housewife'],4))
    ss(72,18,round(p['other_person'],4))

    wb.save(out_path)


def process_files(input_path, out1_path, out_final_path, template_path=None):
    info, data = read_summary5(input_path)
    build_summary1(info, data, out1_path)
    build_summary_final(info, data, out_final_path, template_path=template_path)
