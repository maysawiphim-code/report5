import streamlit as st
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import io, tempfile, os, re, math
import numpy as np
import urllib.request
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import DataBarRule

st.set_page_config(layout="wide")
st.title("📊 ระบบสร้างรายงานสรุปผลอัตโนมัติ")

# ── ดาวน์โหลด Sarabun font ถ้ายังไม่มี (รองรับ Streamlit Cloud) ──
_FONT_PATH = "/tmp/Sarabun-Regular.ttf"
if not os.path.exists(_FONT_PATH):
    urllib.request.urlretrieve(
        "https://github.com/google/fonts/raw/main/ofl/sarabun/Sarabun-Regular.ttf",
        _FONT_PATH
    )
    fm.fontManager.addfont(_FONT_PATH)

def get_thai_font():
    return fm.FontProperties(fname=_FONT_PATH)

thai_font = get_thai_font()

def fmt_pct(v):
    try: return f"{float(v):.2%}"
    except: return str(v)

def fmt_num(v):
    try:
        v = float(v)
        return f"{int(v):,}" if v == int(v) else f"{v:,.1f}"
    except: return str(v)

# ============================================================
# parse_summary2
# ============================================================
def parse_summary2(df):
    r = df.values
    data_col = header_row = None
    for i, row in enumerate(r):
        for j, val in enumerate(row):
            if str(val).strip() == 'ผลัดเช้า':
                data_col = j; header_row = i; break
        if data_col: break
    if data_col is None:
        raise ValueError("ไม่พบคอลัมน์ 'ผลัดเช้า'")

    cols_idx  = [data_col, data_col+1, data_col+2, data_col+3, data_col+4]
    cols_name = ['ผลัดเช้า','ผลัดบ่าย','ผลัดดึก','รวมทั้งวัน','%สัดส่วน']

    def safe(v): return '' if isinstance(v, float) and np.isnan(v) else v

    sections = {k:[] for k in ['เพศ','อายุ','อาชีพ','ทิศคน','ประเภทรถ','ทิศรถ']}
    total_person_row = total_car_row = None
    current_main = current_group = None

    for i in range(header_row+1, len(r)):
        row = r[i]
        col0 = str(safe(row[0])).strip()
        col1 = str(safe(row[1])).strip() if len(row)>1 else ''
        col2 = str(safe(row[2])).strip() if len(row)>2 else ''
        if col0 == 'คน': current_main = 'คน'
        if col0 == 'รถ': current_main = 'รถ'
        if 'จำนวนคน' in col1: total_person_row = i; continue
        if 'จำนวนรถ' in col1: total_car_row = i; continue
        if col1:
            if current_main == 'คน':
                if   'เพศ'    in col1: current_group = 'เพศ'
                elif 'อายุ'   in col1: current_group = 'อายุ'
                elif 'อาชีพ'  in col1: current_group = 'อาชีพ'
                elif 'ทิศ'    in col1: current_group = 'ทิศคน'
            elif current_main == 'รถ':
                if   'ประเภท' in col1: current_group = 'ประเภทรถ'
                elif 'ทิศ'    in col1: current_group = 'ทิศรถ'
        if col2 and current_group and current_main:
            vals = {'หัวข้อ': col2}
            for ci, cn in zip(cols_idx, cols_name):
                try:
                    v = row[ci] if ci < len(row) else ''
                    vals[cn] = '' if (isinstance(v,float) and np.isnan(v)) else v
                except: vals[cn] = ''
            sections[current_group].append(vals)

    def to_df(rows): return pd.DataFrame(rows, columns=['หัวข้อ']+cols_name)
    def tv(idx):
        row = r[idx]
        return [row[ci] if ci < len(row) else 0 for ci in cols_idx[:4]]

    return {
        'เพศ':to_df(sections['เพศ']), 'อายุ':to_df(sections['อายุ']),
        'อาชีพ':to_df(sections['อาชีพ']), 'ทิศคน':to_df(sections['ทิศคน']),
        'ประเภทรถ':to_df(sections['ประเภทรถ']), 'ทิศรถ':to_df(sections['ทิศรถ']),
        'total_person': tv(total_person_row) if total_person_row else [0,0,0,0],
        'total_car':    tv(total_car_row)    if total_car_row    else [0,0,0,0],
    }

# ============================================================
# parse_hourly
# ============================================================
def parse_hourly(df):
    r = df.values
    time_re = re.compile(r'\d{2}:\d{2}\s*[-–]\s*\d{2}:\d{2}')

    hourly_start = person_col = None
    for i, row in enumerate(r):
        for j, val in enumerate(row):
            if time_re.match(str(val).strip()):
                hourly_start = i; person_col = j+1; break
        if hourly_start is not None: break
    if hourly_start is None:
        raise ValueError("ไม่พบข้อมูลรายชั่วโมง")

    CAR_COLS = {
        'รถรวมทั้งหมด': 18, 'รถจักรยาน': 19,
        'รถจักรยานยนต์': 20, 'รถส่วนบุคคล': 21, 'รถอื่นๆ': 22,
    }
    DIR_COLS = {
        'ฝั่งเป้าหมาย': 27, 'ฝั่งตรงข้าม': 30,
        'เข้าซอย': 33, 'ออกซอย': 36,
    }

    def sv(row, col):
        try:
            v = row[col]
            return 0.0 if (isinstance(v,float) and np.isnan(v)) or v is None else float(v)
        except: return 0.0

    hourly = []
    for i in range(hourly_start, len(r)):
        row = r[i]
        time_str = str(row[person_col-1]).strip() if person_col-1 < len(row) else ''
        if not time_re.match(time_str): continue
        entry = {'เวลา': time_str, 'คน': sv(row, person_col)}
        for name, col in {**CAR_COLS, **DIR_COLS}.items():
            entry[name] = sv(row, col)
        hourly.append(entry)

    df_out = pd.DataFrame(hourly)

    n = len(df_out)
    shift_size = 8

    active_dir = [k for k in DIR_COLS if df_out[k].sum() > 0]
    active_car = [k for k in CAR_COLS if df_out[k].sum() > 0]

    shift_totals = {}
    for col in ['คน'] + list(CAR_COLS.keys()) + list(DIR_COLS.keys()):
        if col not in df_out.columns: continue
        vals = df_out[col].values
        shift_totals[col] = {
            'เช้า': float(vals[:shift_size].sum())             if n >= shift_size    else float(vals.sum()),
            'บ่าย': float(vals[shift_size:shift_size*2].sum()) if n >= shift_size*2  else 0.0,
            'ดึก':  float(vals[shift_size*2:].sum())           if n >= shift_size*3  else 0.0,
            'รวม':  float(vals.sum()),
        }

    df_out.attrs['active_dir']    = active_dir
    df_out.attrs['active_car']    = active_car
    df_out.attrs['shift_totals']  = shift_totals
    return df_out

# ============================================================
# Peak detection
# ============================================================
def find_peak_groups(vals, threshold_pct=0.70):
    if len(vals) == 0: return []
    mx = max(vals) if max(vals) > 0 else 1
    in_peak = [v >= mx * threshold_pct for v in vals]
    groups, start = [], None
    for i, v in enumerate(in_peak):
        if v and start is None: start = i
        elif not v and start is not None: groups.append((start, i-1)); start = None
    if start is not None: groups.append((start, len(in_peak)-1))
    return groups

def peak_time_str(groups, times):
    parts = []
    for s, e in groups:
        ts = times[s].split('-')[0].strip()
        te = times[e].split('-')[-1].strip()
        parts.append(f"{ts}-{te}" if s != e else times[s])
    if len(parts) > 1: return ', '.join(parts[:-1]) + f' และ {parts[-1]}'
    return parts[0] if parts else ''

# ============================================================
# make_graph
# ============================================================
def _fmt_val(v):
    return f"{math.ceil(v):,}"

def make_graph(df_h, col, color, title, peak_groups, shift_totals, unit='คน'):
    plt.close('all')
    fig, ax = plt.subplots(figsize=(20, 8), dpi=150)
    x = np.arange(len(df_h))
    vals = df_h[col].values.astype(float)
    mx = max(vals) if max(vals) > 0 else 1
    st_data = shift_totals.get(col, {'เช้า':0,'บ่าย':0,'ดึก':0,'รวม':0})
    total_val = st_data['รวม']

    ax.plot(x, vals, marker='o', markersize=10, color=color, linewidth=2.5, zorder=3)

    for i, val in enumerate(vals):
        label = f"{math.ceil(val):,}"
        kw = {'ha':'center','fontsize':13,'fontweight':'bold','zorder':4,
              'fontproperties': thai_font}
        ax.text(i, val + mx*0.055, label, **kw)

    n = len(df_h)
    shift_size = 8
    for sp in [shift_size-0.5, shift_size*2-0.5]:
        if sp < n:
            ax.axvline(x=sp, color='#555555', linestyle='--', alpha=0.7, linewidth=1.8)

    shift_info = [
        (shift_size/2-0.5,   f"ผลัดเช้า\n{_fmt_val(st_data['เช้า'])} {unit}"),
        (shift_size*1.5-0.5, f"ผลัดบ่าย\n{_fmt_val(st_data['บ่าย'])} {unit}"),
        (shift_size*2.5-0.5, f"ผลัดดึก\n{_fmt_val(st_data['ดึก'])} {unit}"),
    ]
    for xp, label in shift_info:
        if xp < n:
            ax.text(xp, mx*1.33, label,
                    ha='center', va='bottom', fontsize=13, fontweight='bold',
                    color='#222222', fontproperties=thai_font)

    y_bot = -mx*0.08; y_top = mx*1.23
    for s, e in peak_groups:
        ax.add_patch(plt.Rectangle(
            (s-0.45, y_bot), e-s+0.9, y_top-y_bot,
            linewidth=2, edgecolor='#333333', facecolor='#FFD54F',
            alpha=0.38, linestyle='--', zorder=2
        ))

    ax.set_xticks(x)
    ax.set_xticklabels(df_h['เวลา'].tolist(),
                       rotation=90, fontsize=11, fontproperties=thai_font)

    ax.set_title(title, fontsize=18, pad=14, fontweight='bold',
                 fontproperties=thai_font)

    ax.set_ylim(y_bot, mx*1.62)
    ax.set_xlim(-0.7, n-0.3)
    ax.grid(axis='y', alpha=0.25, linestyle=':', linewidth=0.8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.set_facecolor('#FAFBFC')
    fig.patch.set_facecolor('#FFFFFF')

    total_label = f"รวมทั้งหมด\n{_fmt_val(total_val)} {unit}"
    ax.text(1.01, 0.5, total_label,
            transform=ax.transAxes, fontsize=13, va='center', ha='left',
            fontweight='bold', color=color, fontproperties=thai_font,
            bbox=dict(boxstyle='round,pad=0.5', facecolor='white',
                      edgecolor=color, linewidth=1.5, alpha=0.95))

    plt.tight_layout(pad=1.5)
    return fig

# ============================================================
# build_excel
# ============================================================
def build_excel(data, df_h, peak_pg, peak_cg, site_name, date_str,
                fig_p, fig_c, dir_figs, car_figs,
                src_summary2_bytes=None, src_hourly_bytes=None):

    wb = Workbook()
    ws = wb.active
    ws.title = 'รายงานสรุปผล'

    hdr     = Font(bold=True, size=11, name='Arial', color='FFFFFF')
    hdr2    = Font(bold=True, size=10, name='Arial', color='FFFFFF')
    reg     = Font(size=10, name='Arial')
    title_f = Font(bold=True, size=14, name='Arial', color='1F4E79')
    fill_hdr   = PatternFill('solid', start_color='1F4E79')
    fill_total = PatternFill('solid', start_color='D9D9D9')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin   = Side(style='thin')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sc(row, col, val, font=None, fill=None, align=None, border=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:  c.font  = font
        if fill:  c.fill  = fill
        if align: c.alignment = align
        if border: c.border = border
        return c

    def write_table(sr, df_in, show_pct=True):
        cols = ['หัวข้อ','ผลัดเช้า','ผลัดบ่าย','ผลัดดึก','รวมทั้งวัน']
        if show_pct: cols.append('%สัดส่วน')
        for ci, col in enumerate(cols, 2):
            sc(sr, ci, col, font=hdr2, fill=fill_hdr, align=center, border=bdr)
        for ri, (_, row_d) in enumerate(df_in.iterrows()):
            xr = sr+1+ri
            is_total = 'ยอดรวม' in str(row_d.get('หัวข้อ',''))
            for ci, col in enumerate(cols, 2):
                val = row_d.get(col, '')
                is_red = False
                if col == '%สัดส่วน' and val != '':
                    try:
                        pv = float(val)
                        is_red = pv > 0.20
                        val = fmt_pct(pv)
                    except: pass
                elif col not in ['หัวข้อ','%สัดส่วน'] and val != '':
                    try:
                        val = math.ceil(float(val))
                    except: pass
                aln   = left if col == 'หัวข้อ' else center
                cfont = Font(size=10, name='Arial',
                             color='FF0000' if is_red else '000000',
                             bold=is_red or is_total)
                cfill = fill_total if is_total else None
                sc(xr, ci, val, font=cfont, fill=cfill, align=aln, border=bdr)
        for i, cl in enumerate(['C','D','E']):
            rule = DataBarRule(start_type='min', start_value=0,
                               end_type='max', end_value=None,
                               color=['FFC7CE','C6EFCE','FFEB9C'][i], showValue=True)
            ws.conditional_formatting.add(f'{cl}{sr+1}:{cl}{sr+len(df_in)}', rule)
        return sr+1+len(df_in)

    tmp_files = []
    def add_graph(fig, row_idx, w=900, h=420):
        t = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
        fig.savefig(t.name, dpi=150, bbox_inches='tight'); t.close()
        tmp_files.append(t.name)
        img = XLImage(t.name)
        img.width = w; img.height = h
        ws.add_image(img, f'B{row_idx}')
        return row_idx + math.ceil(h / 19)

    shift_totals = df_h.attrs.get('shift_totals', {})
    times_list   = df_h['เวลา'].tolist()
    tp = data['total_person']; tc = data['total_car']
    df_occ  = data['อาชีพ'];   df_dir  = data['ทิศคน']
    df_sex  = data['เพศ'];     df_age  = data['อายุ']
    df_ctyp = data['ประเภทรถ']; df_cdir = data['ทิศรถ']

    top2_occ = df_occ.nlargest(2,'%สัดส่วน')
    occ1, occ2 = top2_occ.iloc[0], top2_occ.iloc[1]
    dir_dom  = df_dir.loc[df_dir['%สัดส่วน'].astype(float).idxmax()]
    top2_age = df_age.nlargest(2,'%สัดส่วน')
    age1, age2 = top2_age.iloc[0], top2_age.iloc[1]
    top2_car = df_ctyp.nlargest(2,'%สัดส่วน')
    car1, car2 = top2_car.iloc[0], top2_car.iloc[1]
    peak_p_str = peak_time_str(peak_pg, times_list)
    peak_c_str = peak_time_str(peak_cg, times_list)

    row = 2
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'รายงานสรุปผลการสำรวจทำเล: {site_name}',font=title_f,align=center); row+=1
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'ข้อมูล ณ {date_str}',font=reg,align=center); row+=2

    # ══ ส่วนคน ══
    total_p = tp[3]
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'🚶 คนเดินผ่าน รวมทั้งวัน {math.ceil(total_p):,} คน',font=hdr,fill=fill_hdr,align=left); row+=1
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'1. กลุ่มอาชีพหลัก: {occ1["หัวข้อ"]} ({fmt_pct(occ1["%สัดส่วน"])}) และ {occ2["หัวข้อ"]} ({fmt_pct(occ2["%สัดส่วน"])})',font=reg,align=left); row+=1
    df_occ_t = pd.concat([df_occ, pd.DataFrame([{'หัวข้อ':'จำนวนคน (ยอดรวม)','ผลัดเช้า':tp[0],'ผลัดบ่าย':tp[1],'ผลัดดึก':tp[2],'รวมทั้งวัน':tp[3],'%สัดส่วน':''}])], ignore_index=True)
    row = write_table(row, df_occ_t) + 2

    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'2. ทิศทางคนส่วนใหญ่: {dir_dom["หัวข้อ"]} ({fmt_pct(dir_dom["%สัดส่วน"])})',font=reg,align=left); row+=1
    dir_f = pd.concat([pd.DataFrame([{'หัวข้อ':'ทิศทางคน (ยอดรวม)','ผลัดเช้า':tp[0],'ผลัดบ่าย':tp[1],'ผลัดดึก':tp[2],'รวมทั้งวัน':tp[3],'%สัดส่วน':1.0}]),df_dir],ignore_index=True)
    row = write_table(row, dir_f) + 2

    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'3. ช่วงเวลาที่คนผ่านหนาแน่น: {peak_p_str} น.',font=reg,align=left); row+=2

    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'4. เพศ/อายุ: ส่วนใหญ่อายุ {age1["หัวข้อ"]} ({fmt_pct(age1["%สัดส่วน"])}) และ {age2["หัวข้อ"]} ({fmt_pct(age2["%สัดส่วน"])})',font=reg,align=left); row+=1
    sa = pd.concat([
        pd.DataFrame([{'หัวข้อ':'เพศ (ยอดรวม)','ผลัดเช้า':tp[0],'ผลัดบ่าย':tp[1],'ผลัดดึก':tp[2],'รวมทั้งวัน':tp[3],'%สัดส่วน':1.0}]),
        df_sex,
        pd.DataFrame([{'หัวข้อ':'ช่วงอายุ','ผลัดเช้า':'','ผลัดบ่าย':'','ผลัดดึก':'','รวมทั้งวัน':'','%สัดส่วน':''}]),
        df_age], ignore_index=True)
    row = write_table(row, sa) + 2

    # ── กราฟคน ──
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,'กราฟคนเดินผ่านรายชั่วโมง',font=hdr,fill=fill_hdr,align=center); row+=1
    row = add_graph(fig_p, row)
    row += 2

    # ══ ส่วนรถ ══
    total_c = tc[3]
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'🚗 รถวิ่งผ่าน รวมทั้งวัน {math.ceil(total_c):,} คัน',font=hdr,fill=fill_hdr,align=left); row+=1
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'1. ประเภทรถหลัก: {car1["หัวข้อ"]} ({fmt_pct(car1["%สัดส่วน"])}) และ {car2["หัวข้อ"]} ({fmt_pct(car2["%สัดส่วน"])})',font=reg,align=left); row+=1
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,'2. ทิศทางรถไม่แตกต่างกัน',font=reg,align=left); row+=1
    car_f = pd.concat([
        df_ctyp,
        pd.DataFrame([{'หัวข้อ':'ทิศทางรถ','ผลัดเช้า':'','ผลัดบ่าย':'','ผลัดดึก':'','รวมทั้งวัน':'','%สัดส่วน':''}]),
        df_cdir,
        pd.DataFrame([{'หัวข้อ':'จำนวนรถ (ยอดรวม)','ผลัดเช้า':tc[0],'ผลัดบ่าย':tc[1],'ผลัดดึก':tc[2],'รวมทั้งวัน':tc[3],'%สัดส่วน':''}]),
    ], ignore_index=True)
    row = write_table(row, car_f) + 2

    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,f'3. ช่วงเวลาที่รถผ่านหนาแน่น: {peak_c_str} น.',font=reg,align=left); row+=2

    # ── กราฟรถ ──
    ws.merge_cells(f'B{row}:L{row}')
    sc(row,2,'กราฟรถวิ่งผ่านรายชั่วโมง',font=hdr,fill=fill_hdr,align=center); row+=1
    row = add_graph(fig_c, row)
    for fig in dir_figs.values():
        row = add_graph(fig, row)
    for fig in car_figs.values():
        row = add_graph(fig, row)

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 32
    for c in 'CDEFG': ws.column_dimensions[c].width = 12
    for c in 'HIJKL': ws.column_dimensions[c].width = 10

    # ── Sheet ต้นฉบับ ──
    for src_bytes, sheet_name in [
        (src_summary2_bytes, 'ข้อมูลสรุป2'),
        (src_hourly_bytes,   'ข้อมูลรายชั่วโมง'),
    ]:
        if not src_bytes:
            continue
        try:
            src_wb = load_workbook(io.BytesIO(src_bytes))
            src_ws = src_wb.active
            new_ws = wb.create_sheet(title=sheet_name)

            pct_column = None
            for row_s in src_ws.iter_rows(min_row=1, max_row=10, values_only=True):
                for idx, cell in enumerate(row_s, 1):
                    if cell and isinstance(cell, str) and 'สัดส่วน' in str(cell):
                        pct_column = idx
                        break
                if pct_column: break

            for ri, src_row in enumerate(src_ws.iter_rows(values_only=True), start=1):
                for ci, val in enumerate(src_row, start=1):
                    c = new_ws.cell(row=ri, column=ci)
                    if isinstance(val, float):
                        if np.isnan(val):
                            c.value = 0
                            c.number_format = '#,##0'
                            continue
                        if sheet_name == 'ข้อมูลรายชั่วโมง':
                            if ri <= 35 and 0 < val < 1.0:
                                c.value = round(val * 100, 4)
                                c.number_format = '0.00"%";-0.00"%";0.00"%";@'
                            elif ri == 71 or (ri > 71 and 0 < val < 1.0):
                                c.value = val
                                c.number_format = '0.0000'
                            else:
                                c.value = math.ceil(val) if val % 1 >= 0.5 else int(val)
                                c.number_format = '#,##0'
                        else:
                            if ci in [3, 4, 5, 6]:
                                if val % 1 >= 0.5:
                                    c.value = math.ceil(val)
                                else:
                                    c.value = int(val)
                                c.number_format = '#,##0'
                            elif (pct_column and ci == pct_column) or (0 < val < 1.0 and ci >= 6):
                                c.value = round(val * 100, 4)
                                c.number_format = '0.00"%";-0.00"%";0.00"%";@'
                            else:
                                c.value = int(round(val))
                                c.number_format = '#,##0'
                    else:
                        c.value = val if val is not None else ''
        except Exception as e:
            st.warning(f"คัดลอก sheet {sheet_name} ไม่สมบูรณ์: {e}")

    buf = io.BytesIO()
    wb.save(buf)
    for f in tmp_files:
        try: os.unlink(f)
        except: pass
    return buf.getvalue()

# ============================================================
# UI
# ============================================================
file2     = st.file_uploader("อัปโหลด สรุป2.xlsx", type=["xlsx","csv"])
file_hour = st.file_uploader("อัปโหลด สรุปรายชั่วโมง.xlsx", type=["xlsx","csv"])
site_name = st.text_input("ชื่อทำเล", value="ทำเลเป้าหมาย")
date_str  = st.text_input("วันที่สำรวจ", value="วันที่ 6 และ 7 มีนาคม 2569")

if file2 and file_hour and st.button("🔍 วิเคราะห์และสร้างรายงาน"):
    try:
        df2  = pd.read_excel(file2,     header=None)
        dfh  = pd.read_excel(file_hour, header=None)
        data = parse_summary2(df2)
        df_h = parse_hourly(dfh)

        shift_totals = df_h.attrs['shift_totals']
        active_dir   = df_h.attrs['active_dir']
        active_car   = df_h.attrs['active_car']
        car_types    = [c for c in active_car if c != 'รถรวมทั้งหมด']

        p_total = math.ceil(float(data['total_person'][3]))
        c_total = math.ceil(float(data['total_car'][3]))
        st.success(f"✅ อ่านข้อมูลสำเร็จ — คนรวม {p_total:,} คน | รถรวม {c_total:,} คัน")

        peak_pg = find_peak_groups(df_h['คน'].values)
        peak_cg = find_peak_groups(df_h['รถรวมทั้งหมด'].values)

        col_colors = {
            'คน':'red', 'รถรวมทั้งหมด':'darkorange',
            'ฝั่งเป้าหมาย':'#E53935', 'ฝั่งตรงข้าม':'#1E88E5',
            'เข้าซอย':'#43A047', 'ออกซอย':'#FB8C00',
            'รถจักรยาน':'#6D4C41', 'รถจักรยานยนต์':'#00897B',
            'รถส่วนบุคคล':'#5E35B1', 'รถอื่นๆ':'#757575',
        }

        fig_p = make_graph(df_h,'คน','red','กราฟแสดงจำนวนคนเดินผ่านรายชั่วโมง',peak_pg,shift_totals,'คน')
        fig_c = make_graph(df_h,'รถรวมทั้งหมด','darkorange','กราฟแสดงจำนวนรถวิ่งผ่านรายชั่วโมง (รวม)',peak_cg,shift_totals,'คัน')

        dir_figs = {n: make_graph(df_h, n, col_colors.get(n,'gray'), f'กราฟทิศทางรถ: {n}',
                                  find_peak_groups(df_h[n].values), shift_totals, 'คัน')
                    for n in active_dir}
        car_figs = {n: make_graph(df_h, n, col_colors.get(n,'gray'), f'กราฟประเภทรถ: {n}',
                                  find_peak_groups(df_h[n].values), shift_totals, 'คัน')
                    for n in car_types}

        st.subheader("📈 กราฟจำนวนคนและรถรวม")
        c1, c2 = st.columns(2)
        with c1: st.pyplot(fig_p)
        with c2: st.pyplot(fig_c)

        if dir_figs:
            st.subheader("🚗 กราฟทิศทางรถ")
            cols = st.columns(min(2, len(dir_figs)))
            for i, (name, fig) in enumerate(dir_figs.items()):
                with cols[i % len(cols)]: st.pyplot(fig)

        file2.seek(0);     src2_bytes = file2.read()
        file_hour.seek(0); srch_bytes = file_hour.read()

        excel_bytes = build_excel(
            data, df_h, peak_pg, peak_cg, site_name, date_str,
            fig_p, fig_c, dir_figs, car_figs,
            src_summary2_bytes=src2_bytes,
            src_hourly_bytes=srch_bytes,
        )
        st.download_button("📥 ดาวน์โหลดรายงานฉบับสมบูรณ์", excel_bytes,
                           "รายงานสรุปผล.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"❌ เกิดข้อผิดพลาด: {e}")
        import traceback; st.code(traceback.format_exc())
